
"""
27卒 6社就職比較表（極東開発工業除く・並び順更新）
- A4横1枚 PDF
- 工場・研究・開発拠点：関西エリア拠点は事業内容サマリー付き
- データ出典: マイナビ2027 / OpenWork / 公式採用HP / kabutan / indeed / 日経
"""
import os, shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
JP = 'HeiseiKakuGo-W5'

OUT_DIR   = r"\\pc7075\380USER_Sec\380BUSEI_G5_Sec\10.中山\ダウンロード"
OUT_XLSX  = os.path.join(OUT_DIR, "27卒_6社就職比較表.xlsx")
TMP_PDF   = r"C:\Users\ka94654\AppData\Local\Temp\cmp6_measure_tmp.pdf"
OUT_LOCAL = r"C:\Users\ka94654\AppData\Local\Temp\27卒_6社就職比較表.pdf"
OUT_PDF   = os.path.join(OUT_DIR, "27卒_6社就職比較表.pdf")

SHEET_MAIN = "6社比較"
SHEET_URLS = "参照URL"

# 並び順: 伯東・ダイニック・ヤマトプロテック・ラサ工業・阪神内燃機工業・SWCC
CN = ["伯東", "ダイニック", "ヤマトプロテック",
      "ラサ工業(株)", "阪神内燃機工業(株)", "SWCC(株)"]
N = len(CN)

# ============================================================
# データ定義
# ============================================================

ROWS = [
    ("【基本情報】", None),
    ("正式社名", [
        "伯東(株)", "ダイニック(株)", "ヤマトプロテック(株)",
        "ラサ工業(株)", "阪神内燃機工業(株)", "SWCC(株)"]),
    ("上場区分", [
        "東証プライム(7433)", "東証スタンダード(3551)", "非上場",
        "東証プライム(4022)", "東証スタンダード(6018)", "東証プライム(5805)"]),
    ("資本金", [
        "81億円(2025/3)", "57億9,565万円", "9,900万円",
        "84億4,300万円", "8億2,905万円(2025/3)", "242億2,100万円"]),
    ("本社所在地", [
        "東京都新宿区(2027年秋四谷移転予定)", "東京都港区(新橋)", "東京都港区(白金台)",
        "東京都千代田区(秋葉原ダイビル)", "兵庫県神戸市中央区(神港ビル)", "神奈川県川崎市川崎区"]),
    ("従業員数", [
        "単体723名/連結1,318名", "単体606名/連結1,089名", "381名(グループ連結744名)",
        "単体455名/連結628名(2025/3)", "302名(2026/3)", "単体1,453名/連結4,945名(2025/3)"]),
    ("創業・歴史", [
        "1953年創業(2023年70周年)", "1919年創業(100年超)", "1918年創業(109年)",
        "1913年創業(113年・旧ラサ島燐砿)", "1918年創業(108年・内航船エンジン55%シェア)", "1936年創業(旧昭和電線・89年)"]),
    ("", None),
    ("【財務・業績】", None),
    ("最新売上高", [
        "連結1,831億円(2025/3期)", "連結440.7億円(2025/3期)", "単体360億円/連結383.6億円(2025/12期)",
        "連結477億2,700万円(2026/3期)", "140億2,800万円(2026/3期)", "連結2,777億3,600万円(2026/3期)"]),
    ("最新営業利益", [
        "連結79億円(2025/3期)", "連結21.4億円(+72.6%/2025/3期)", "非公表",
        "営業60億1,200万円/経常61億9,100万円(2026/3期)", "8億2,400万円(2026/3期/+34.7%)", "連結273億2,000万円(2026/3期/+30.5%)"]),
    ("過去10年の業績トレンド", [
        "中長期で拡大・2026/3期は一時減益(ケミカル低迷)", "売上横ばい〜緩増、近年利益急回復", "350〜380億前後で緩やか成長・新工場建設中",
        "半導体向けリン酸で急成長・2026/3期も増収増益継続", "内航船需要安定・2026/3期増収増益(+34.7%)", "4年連続増収・電力インフラ/DC需要で急拡大"]),
    ("株価推移(10年)", [
        "プライム・中長期上昇基調", "横ばい〜小幅上昇", "該当なし(非上場)",
        "プライム(4022)・半導体需要で急上昇", "スタンダード(6018)・横ばい〜緩増", "プライム(5805)・近年大幅上昇(電力/DC恩恵)"]),
    ("", None),
    ("【勤務地・拠点(営業所除く)】", None),
    ("理系・技術の主な勤務地", [
        "本社(新宿)・技術センター(伊勢原)・四日市研究所/工場", "埼玉・滋賀・王子・富士・真岡工場、本社(東京)", "中央研究所・東京工場(茨城)、大阪工場(堺)",
        "本社(東京)・岩手・宮城・群馬・大阪・福岡各工場", "本社(神戸)・明石工場・玉津工場・播磨工場", "本社(川崎)・三重事業所・愛知工場・茨城・仙台ほか"]),
    ("工場・研究・開発拠点", [
        "四日市研究所/工場・技術センター(伊勢原)\n【関西】大阪支社(淀川区・半導体デバイス技術営業/工業薬品商社機能)",
        "埼玉・真岡(栃木)・富士(静岡)・尾道(広島)工場\n【関西】滋賀工場(犬上郡多賀町・ブッククロス/壁紙/有機EL用水分除去シート等製造・最大拠点)",
        "茨城工場・東京R&Dセンター\n【関西】大阪工場(堺市美原区・消火器/消火装置製造・中央研究所で世界初K/SMOKE開発)",
        "宮古(岩手)・三本木(宮城)・伊勢崎(群馬)・羽犬塚(福岡)工場\n【関西】大阪工場(大阪市大正区・高純度燐酸/水処理薬剤/表面処理剤製造・売上の約7割を担う主力工場)",
        "【関西のみ・全拠点が関西圏】\n本社(神戸市)・明石工場(明石市)・玉津工場(神戸市西区)・播磨工場(播磨町)：舶用ディーゼルエンジン/スラスター/推進装置の設計・製造・整備",
        "茨城工場・仙台事業所・愛知工場\n【関西近接】三重事業所(いなべ市北勢町・裸銅線/被覆線/巻線製造)・大阪支社(電線/光ファイバ営業)"]),
    ("27卒主な募集職種", [
        "技術営業・研究・技術・化学営業・営業事務", "開発技術職・生産設備技術職・営業職・経理職", "技術系総合職(研究・生産・設計・施工・保守)/総合職(文理不問)",
        "技術職(開発・製造・品質)・営業・管理(11〜15名)", "技術系(エンジン設計・開発・製造)・営業系・事務系", "技術系総合職/エリア総合職(研究・生産・営業等)"]),
    ("", None),
    ("【OpenWork・働き方】", None),
    ("OpenWork総合評価(5点)", [
        "3.38(109人)", "2.70(18人)", "2.51(口コミ2件・参考値)",
        "2.92(13人)", "3.05(9人)", "2.89(86人)"]),
    ("待遇面の満足度", [
        "3.4", "2.6", "2.75",
        "3.1", "3.0", "2.9"]),
    ("社員の士気", [
        "2.9", "2.6", "2.7",
        "2.7", "2.8", "2.5"]),
    ("20代成長環境", [
        "2.9", "2.2", "—",
        "2.5", "2.8", "2.7"]),
    ("平均年収(会社全体)", [
        "日経945万/OW626万(43人)", "日経589万/OW466万", "非公表(indeed参考:職種別50万〜)",
        "日経650万/OW506万(10人)", "日経624万/OW非公表(9人)", "日経729万/OW530万(36人)"]),
    ("30歳前後の年収目安", [
        "非公表(商社系・平均年収水準は高め)", "450〜500万円台", "参考:indeed製造職31.9万/月",
        "非公表(25-29歳参考値要確認)", "非公表(勤続長・年功序列)", "25-29歳平均約433万円(OW参考)"]),
    ("有給休暇消化率", [
        "69.0%(OW)/実績14.8日", "58.8%", "低い評価あり/実績8.4日",
        "69.5%(OW)/実績14.3日(2025年度)", "77.2%(OW)/実績14.3日(2025年度)", "57.5%(OW)/実績15.5日(2023年度)"]),
    ("残業時間(月・マイナビ/公式)", [
        "7.3時間", "10時間程度(2023年度・公式HP)", "18.7時間(2025年・公式HP)",
        "7.6時間(2025年度)", "22.2時間(2025年度)", "20時間以内"]),
    ("残業時間(月・OpenWork)", [
        "23.2時間", "12.2時間", "17.7時間",
        "7.2時間", "8.2時間", "21.9時間"]),
    ("平均勤続年数", [
        "13年", "17年", "11.6年",
        "19.0年(2025/3)", "19.6年(男19.1女21.4)", "16.8年"]),
    ("", None),
    ("【給与・福利厚生】", None),
    ("初任給(大卒・月)", [
        "293,000〜303,000円(理系・首都圏・手当含む)", "237,000円", "251,000円(2025/11引上げ後)",
        "266,500円(大卒・手当込)※院了287,100円", "234,000円(大卒・2025/4)※院了252,000円", "287,000円(大卒総合職)※院了309,000円"]),
    ("賞与実績(昨年・年間)", [
        "年2回・口コミ参考約5ヶ月(要確認)", "年2回・昨年実績約4ヶ月分", "年2回・7.2ヶ月分",
        "年2回・5.81ヶ月(2026年実績)", "年2回(7月・12月)", "年2回(6月・12月)"]),
    ("年間休日", [
        "125日前後(採用サイトFAQ)", "120日前後(本社)", "127日(公式HP)",
        "123日", "125日(公式HP)", "121日以上(公式HP)"]),
    ("住宅手当・社宅", [
        "社宅・借上社宅(勤務地により)", "独身寮(月1万程度)", "借上社宅・住宅手当",
        "借上げ社宅(東京基準：月6万円補助)", "独身寮・住宅資金貸付制度", "社宅(三重・愛知工場)・住宅手当"]),
    ("食事補助", [
        "社員食堂(工場)", "食事手当", "食事補助(公式HP記載)/スポーツジム/保養所",
        "要確認", "要確認", "要確認"]),
    ("資格取得支援", [
        "TOEIC支援・英語研修・キャリアアップ支援", "資格・通信教育補助", "消防設備士受験料全額+奨励金",
        "資格奨励金・通信教育補助(70%負担)", "英会話研修補助・各種階級別研修", "各種研修・eラーニング・ジョブチャレンジ"]),
    ("その他特色制度", [
        "メンター・1on1・健康経営優良法人2026・2030ビジョン", "在宅勤務・育児短時間・ジョブリターン・65歳継続雇用", "水・木ノー残業・19時PC停止・誕生日休暇",
        "フレックス・社員持株会・退職年金", "フレックス・時間単位年休・テレワーク可", "フレックス・テレワーク・ジョブチャレンジ制度"]),
    ("育児休業取得率(男性)", [
        "64.7%(2024年度)", "55.6%", "100%",
        "75.0%(2025年度)", "62.5%(2025年度)", "28.6%(2023年度)"]),
    ("", None),
    ("【27卒・就活メモ】", None),
    ("事業の軸", [
        "エレクトロニクス商社×ケミカルメーカー(半導体・工業薬品)", "化学コーティング・フィルム・クロス", "防災・消防設備(国内トップシェア)",
        "化成品(高純度リン酸/半導体向け)・機械・電子材料", "舶用ディーゼルエンジン・推進装置(内航船55%シェア)", "電線・ケーブル・光ファイバ・超電導・電力インフラ・DC向け"]),
    ("マイナビ2027", [
        "掲載あり(corp290)・オンライン説明会・エントリー受付中", "掲載あり(corp57404)・4職種募集", "掲載あり(corp58621)・総合職/技術系総合職",
        "掲載あり(corp57519)・募集11〜15名", "掲載あり(corp55663)・募集6〜10名", "掲載あり(corp85245)"]),
    ("採用実績に大阪電気通信大学(マイナビ2027)", [
        "なし(電気通信大学はあり)", "あり(大学・採用実績学校)", "あり(大学)",
        "なし(掲載なし)", "あり(大学・採用実績校に明記)", "あり(2025年度採用実績に記載)"]),
    ("データ出典・更新", [
        "マイナビ2027/OpenWork/公式HP/kabutan/indeed(2026年6月)", "同上", "同上",
        "同上", "同上", "同上"]),
]

# ざっくり比較
AXIS_ROWS = [
    ("【ざっくり比較（就活の軸）】", None),
    ("初任給（大卒・月給）",
     "伯東(29.3〜30.3万・手当込) ＞ SWCC(28.7万) ＞ ラサ工業(26.65万) ＞ ヤマトプロテック(25.1万) ＞ ダイニック(23.7万) ＞ 阪神内燃(23.4万)"),
    ("資本金",
     "SWCC(242億) ＞ ラサ工業(84億) ＞ 伯東(81億) ＞ ダイニック(57億) ＞ 阪神内燃・ヤマトプロテック(〜9億)"),
    ("会社規模・売上",
     "SWCC(2,777億) ＞ 伯東(1,831億) ＞ ラサ工業(477億) ＞ ダイニック(441億) ＞ ヤマトプロテック(384億) ＞ 阪神内燃(140億)"),
    ("口コミ総合（OpenWork）",
     "伯東(3.38) ＞ 阪神内燃(3.05) ＞ ラサ工業(2.92) ＞ SWCC(2.89) ＞ ダイニック(2.70) ＞ ヤマトプロテック(2.51)"),
    ("残業の少なさ（マイナビ/公式実績）",
     "ラサ工業(7.6h) ＞ 伯東(7.3h) ＞ ダイニック(10h程度) ＞ ヤマトプロテック(18.7h) ＞ SWCC(20h以内) ＞ 阪神内燃(22.2h)"),
    ("賞与（昨年実績・年間）",
     "ヤマトプロテック(7.2ヶ月) ＞ ラサ工業(5.81ヶ月/2026年実績) ＞ 伯東・ダイニック(4〜5ヶ月前後) ／ SWCC・阪神内燃は年2回(実績月数要確認)"),
    ("理系・化学系の専門性",
     "ラサ工業(半導体向けリン酸・化学急成長)／伯東(商社×化学)／ダイニック(コーティング)／SWCC(電線・超電導・DC向け急拡大)／ヤマト(防災)／阪神内燃(舶用エンジン)"),
    ("関西エリアとの関わり",
     "阪神内燃機工業(全拠点が神戸・明石・播磨)◎ ／ ラサ工業(大阪工場が主力・売上7割) ／ ヤマトプロテック(大阪工場・研究開発) ／ ダイニック(滋賀工場が最大拠点) ／ SWCC(三重事業所) ／ 伯東(大阪支社)"),
    ("社会貢献・インフラ感",
     "SWCC(電力インフラ・DC)・ヤマトプロテック(防災)・阪神内燃(海運インフラ) ＞ 伯東・ラサ工業(半導体) ＞ ダイニック"),
    ("上場・将来性（株価）",
     "プライム: 伯東・ラサ工業(半導体需要◎)・SWCC(電力/DC急成長◎) ／ スタンダード: ダイニック・阪神内燃 ／ 非上場: ヤマトプロテック"),
    ("", None),
    ("出典: マイナビ2027 / OpenWork / 公式採用HP / kabutan・irbank(業績) / indeed / 日経会社情報。2026年6月時点。",
     "賞与・初任給の一部はユーザー確認値を反映。山陽色素・極東開発工業は今回リストより除外。"),
]

URLS = [
    ("伯東",           "https://job.mynavi.jp/27/pc/search/corp290/outline.html"),
    ("ダイニック",     "https://job.mynavi.jp/27/pc/search/corp57404/outline.html"),
    ("ヤマトプロテック","https://job.mynavi.jp/27/pc/search/corp58621/outline.html"),
    ("ラサ工業",       "https://job.mynavi.jp/27/pc/search/corp57519/outline.html"),
    ("阪神内燃機工業", "https://job.mynavi.jp/27/pc/search/corp55663/outline.html"),
    ("SWCC",           "https://job.mynavi.jp/27/pc/search/corp85245/outline.html"),
]

# ============================================================
# Excel 生成
# ============================================================
HC  = "1F4E79"
SC  = "2E75B6"
ALT = "DEEAF1"
W   = "FFFFFF"

def border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

def write_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_MAIN
    ws.column_dimensions['A'].width = 26
    for i in range(2, 2 + N):
        ws.column_dimensions[get_column_letter(i)].width = 27

    ws.row_dimensions[1].height = 26
    for col, val in enumerate(["比較項目"] + CN, 1):
        c = ws.cell(row=1, column=col, value=val)
        c.font = Font(name='Meiryo', bold=True, color=W, size=10)
        c.fill = PatternFill("solid", fgColor=HC)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border()

    all_rows = ROWS + AXIS_ROWS
    ridx = 2
    alt = False
    for label, values in all_rows:
        is_sec  = label.startswith('【')
        is_emp  = label == ''
        is_axis = isinstance(values, str)
        is_foot = is_axis and label.startswith('出典')

        if is_sec:
            bg = SC
        elif is_emp:
            bg = W
        else:
            bg = ALT if alt else W
            alt = not alt

        # 工場・研究・開発拠点は行を高めに
        is_kansai_row = label == "工場・研究・開発拠点"
        ws.row_dimensions[ridx].height = 54 if is_kansai_row else (36 if (is_sec or is_axis) else 30)

        a = ws.cell(row=ridx, column=1, value=label)
        a.font = Font(name='Meiryo', bold=is_sec or is_foot, color=(W if is_sec else "000000"), size=9)
        a.fill = PatternFill("solid", fgColor=bg)
        a.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        a.border = border()

        if values is None:
            for col in range(2, 2 + N):
                c = ws.cell(row=ridx, column=col, value="")
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = border()
        elif is_axis:
            ws.merge_cells(start_row=ridx, start_column=2, end_row=ridx, end_column=1 + N)
            c = ws.cell(row=ridx, column=2, value=values)
            c.font = Font(name='Meiryo', size=8.5)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = border()
        else:
            for i, val in enumerate(values):
                c = ws.cell(row=ridx, column=i + 2, value=val)
                c.font = Font(name='Meiryo', size=8.5)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal='left' if is_kansai_row else 'center',
                                        vertical='center', wrap_text=True)
                c.border = border()
        ridx += 1

    ws2 = wb.create_sheet(SHEET_URLS)
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 80
    ws2.append(["会社名", "マイナビ2027 URL"])
    for c in ws2[1]:
        c.font = Font(bold=True, color=W)
        c.fill = PatternFill("solid", fgColor=HC)
        c.alignment = Alignment(horizontal='center')
    for row in URLS:
        ws2.append(list(row))

    wb.save(OUT_XLSX)
    print(f"Excel saved: {OUT_XLSX}")

# ============================================================
# PDF 生成
# ============================================================
def jp(text, fs=4.8, col=colors.black, align='CENTER'):
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    return Paragraph(str(text) if text else '', ParagraphStyle(
        'j', fontName=JP, fontSize=fs, leading=fs * 1.35,
        textColor=col,
        alignment=TA_LEFT if align == 'LEFT' else TA_CENTER,
        wordWrap='CJK',
        spaceAfter=0, spaceBefore=0))

def build_table(fs_hdr=5.5, fs_sec=5.0, fs_lbl=4.6, fs_val=4.3, fs_ax=4.3):
    PW, PH = landscape(A4)
    ML = MR = 8
    cw = PW - ML - MR
    LW = 82
    DW = (cw - LW) / N

    HC_c  = colors.HexColor('#1F4E79')
    SC_c  = colors.HexColor('#2E75B6')
    ALT_c = colors.HexColor('#DEEAF1')
    W_c   = colors.white
    BK    = colors.black

    table_data = []
    row_styles = []
    span_cmds  = []

    table_data.append(
        [jp("比較項目", fs=fs_hdr, col=W_c, align='LEFT')] +
        [jp(c, fs=fs_hdr, col=W_c) for c in CN]
    )
    row_styles.append(('BACKGROUND', (0,0), (-1,0), HC_c))

    ridx = 1
    alt  = False
    for label, values in ROWS + AXIS_ROWS:
        is_sec  = label.startswith('【')
        is_emp  = label == ''
        is_axis = isinstance(values, str)
        is_kansai_row = label == "工場・研究・開発拠点"

        if is_sec:
            bg = SC_c
        elif is_emp:
            bg = W_c
        else:
            bg = ALT_c if alt else W_c
            alt = not alt

        lc = W_c if is_sec else BK

        if values is None:
            row = [jp(label, fs=fs_sec if is_sec else 3.5, col=lc, align='LEFT')] + \
                  [jp('') for _ in CN]
            if not is_emp:
                span_cmds.append(('SPAN', (0, ridx), (N, ridx)))
        elif is_axis:
            row = [jp(label, fs=fs_lbl, col=BK, align='LEFT'),
                   jp(values, fs=fs_ax, col=BK, align='LEFT')] + \
                  [jp('') for _ in range(N-1)]
            span_cmds.append(('SPAN', (1, ridx), (N, ridx)))
        else:
            fs_use = fs_val * 0.9 if is_kansai_row else fs_val
            row = [jp(label, fs=fs_sec if is_sec else fs_lbl, col=lc, align='LEFT')] + \
                  [jp(v, fs=fs_use, col=(W_c if is_sec else BK), align='LEFT' if is_kansai_row else 'CENTER')
                   for v in values]

        table_data.append(row)
        row_styles.append(('BACKGROUND', (0, ridx), (-1, ridx), bg))
        if is_sec:
            row_styles.append(('TEXTCOLOR', (0, ridx), (-1, ridx), W_c))
        ridx += 1

    col_widths = [LW] + [DW] * N
    base = [
        ('FONTNAME',      (0,0),(-1,-1), JP),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
        ('ALIGN',         (0,0),(0,-1),  'LEFT'),
        ('GRID',          (0,0),(-1,-1), 0.25, colors.HexColor('#AAAAAA')),
        ('LEFTPADDING',   (0,0),(-1,-1), 2),
        ('RIGHTPADDING',  (0,0),(-1,-1), 2),
        ('TOPPADDING',    (0,0),(-1,-1), 2),
        ('BOTTOMPADDING', (0,0),(-1,-1), 2),
    ]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(base + row_styles + span_cmds))
    return t, col_widths

def write_pdf():
    from reportlab.pdfgen import canvas as cv_mod

    PW, PH = landscape(A4)
    ML = MR = 8; MT = MB = 10
    avail_w = PW - ML - MR
    avail_h = PH - MT - MB

    t, col_widths = build_table()

    tmp_cv = cv_mod.Canvas(TMP_PDF, pagesize=landscape(A4))
    nat_w, nat_h = t.wrapOn(tmp_cv, avail_w, 99999)
    print(f"  Table natural size: {nat_w:.1f} x {nat_h:.1f}  avail: {avail_w:.1f} x {avail_h:.1f}")

    scale = min(1.0, avail_h / nat_h, avail_w / nat_w)
    print(f"  Scale: {scale:.4f}")

    c = cv_mod.Canvas(OUT_LOCAL, pagesize=landscape(A4))
    c.saveState()
    c.translate(ML, MB + avail_h - nat_h * scale)
    c.scale(scale, scale)
    t.drawOn(c, 0, 0)
    c.restoreState()
    c.save()
    print(f"PDF saved locally: {OUT_LOCAL}")

    try:
        shutil.copy2(OUT_LOCAL, OUT_PDF)
        print(f"PDF copied to: {OUT_PDF}")
    except Exception as e:
        print(f"[WARN] network copy failed: {e}")
        print(f"  → ローカルファイルを手動でコピーしてください: {OUT_LOCAL}")

if __name__ == '__main__':
    write_xlsx()
    write_pdf()
    print("Done.")
