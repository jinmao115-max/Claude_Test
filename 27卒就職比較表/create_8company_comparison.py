
"""
27卒 8社就職比較表 生成スクリプト
- A4横1枚 PDF
- 元7社PDF と同じ行構成（営業拠点(参考)なし）
- ざっくり比較はセルを結合して横幅フル表示
"""
import os, shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
JP = 'HeiseiKakuGo-W5'

OUT_DIR  = r"\\pc7075\380USER_Sec\380BUSEI_G5_Sec\10.中山\ダウンロード"
OUT_XLSX = os.path.join(OUT_DIR, "27卒_8社就職比較表.xlsx")
TMP_PDF  = r"C:\Users\ka94654\AppData\Local\Temp\27卒_8社就職比較表_tmp.pdf"
OUT_PDF  = os.path.join(OUT_DIR, "27卒_8社就職比較表.pdf")

SHEET_MAIN = "8社比較"
SHEET_URLS = "参照URL"

CN = ["伯東", "ダイニック", "ヤマトプロテック", "山陽色素",
      "阪神内燃機工業(株)", "ラサ工業(株)", "SWCC(株)", "極東開発工業(株)"]
N = len(CN)

# ============================================================
# データ定義
# ============================================================
# 通常行: (ラベル, [企業1値, ..., 企業8値])
# ざっくり比較行: (ラベル, 値テキスト)  ← 文字列1つ
# ============================================================

ROWS = [
    ("【基本情報】", None),
    ("正式社名", [
        "伯東(株)", "ダイニック(株)", "ヤマトプロテック(株)", "山陽色素(株)",
        "阪神内燃機工業(株)", "ラサ工業(株)", "SWCC(株)", "極東開発工業(株)"]),
    ("上場区分", [
        "東証プライム(7433)", "東証スタンダード(3551)", "非上場", "非上場",
        "東証スタンダード(6018)", "東証プライム(4022)", "東証プライム(5805)", "東証プライム(7226)"]),
    ("資本金", [
        "81億円(2025/3)", "57億9,565万円", "9,900万円", "9,100万円",
        "8億2,905万円(2025/3)", "84億4,300万円", "242億2,100万円", "118億9,986万円"]),
    ("本社所在地", [
        "東京都新宿区(2027年秋四谷移転予定)", "東京都港区(新橋)", "東京都港区(白金台)", "兵庫県姫路市",
        "兵庫県神戸市中央区(神港ビル)", "東京都千代田区(秋葉原ダイビル)", "神奈川県川崎市川崎区", "大阪府大阪市中央区淡路町"]),
    ("従業員数", [
        "単体723名/連結1,318名", "単体606名/連結1,089名", "381名(グループ連結744名)", "365名",
        "302名(2026/3)", "単体455名/連結628名(2025/3)", "単体1,453名/連結4,945名(2025/3)", "単独1,180名/連結3,481名(2025/3)"]),
    ("創業・歴史", [
        "1953年創業(2023年70周年)", "1919年創業(100年超)", "1918年創業(109年)", "1931年創業(有機顔料専業)",
        "1918年創業(108年・内航船エンジン55%シェア)", "1913年創業(113年・旧ラサ島燐砿)", "1936年創業(旧昭和電線・89年)", "1955年設立(71年)"]),
    ("", None),
    ("【財務・業績】", None),
    ("最新売上高", [
        "連結1,831億円(2025/3期)", "連結440.7億円(2025/3期)", "360.0億円(2025/12期)", "240億円(2025/3期)",
        "140億2,800万円(2026/3期)", "連結477億2,700万円(2026/3期)", "連結2,378億6,200万円(2025/3期)", "連結1,404億4,900万円(2025/3期)"]),
    ("最新営業利益", [
        "連結79億円(2025/3期)", "連結21.4億円(+72.6%)", "非公表", "非公表(当期純利益約19億円・25/3期)",
        "非公表", "経常61億9,100万円(2026/3期)", "連結209億3,500万円(2025/3期)", "非公表"]),
    ("過去10年の業績トレンド", [
        "売上・利益とも中長期で拡大(商社×メーカー)", "売上横ばい〜緩増、近年利益急回復", "350億前後で緩やか増", "純利益は変動あり、23期以降は大幅改善",
        "内航船需要安定・緩増基調", "半導体向けリン酸で拡大基調・近年急成長", "4年連続増収・好調", "2年連続増収(特装車需要堅調)"]),
    ("株価推移(10年)", [
        "プライム上場・中長期上昇基調", "横ばい〜小幅上昇", "該当なし(非上場)", "該当なし(非上場)",
        "スタンダード(6018)・横ばい〜緩増", "プライム(4022)・半導体需要で上昇基調", "プライム(5805)・近年大幅上昇", "プライム(7226)・要確認"]),
    ("", None),
    ("【勤務地・拠点(営業所除く)】", None),
    ("理系・技術の主な勤務地", [
        "本社(新宿)・技術センター(伊勢原)・四日市研究所/工場", "埼玉・滋賀・王子・富士・真岡工場、本社(東京)", "中央研究所・東京工場(茨城)、大阪工場(堺)", "姫路工場・東海工場(静岡)、研究開発部門",
        "本社(神戸)・明石工場・玉津工場・播磨工場", "本社(東京)・岩手・宮城・群馬・大阪・福岡各工場", "本社(川崎)・三重事業所・愛知工場・茨城・兵庫ほか", "本社(大阪)・大和工場(神奈川)・小牧工場(愛知)・三木工場(兵庫)"]),
    ("工場・研究拠点(例)", [
        "四日市研究所/工場、技術センター(伊勢原)", "八日市・鈴鹿・尾道・長野ほか", "東京工場(茨城)、大阪工場(堺)", "姫路工場(本社併設)、東海工場(掛川)",
        "明石・玉津(神戸市西区)・播磨(播磨町)工場", "宮古(岩手)・三本木(宮城)・伊勢崎(群馬)・大阪・羽犬塚(福岡)工場",
        "三重事業所・愛知工場・茨城・仙台・大阪・九州ほか", "大和(神奈川)・小牧(愛知)・三木(兵庫)・飯塚(福岡)工場"]),
    ("27卒主な募集職種", [
        "技術営業・研究・技術・化学営業・営業事務", "技術・研究開発・製造など", "技術系総合職(研究・生産・設計・施工・保守)", "技術系総合職(化学)・工場設備職・営業・総合事務",
        "技術系(エンジン設計・開発・製造)・事務系", "技術系・事務系総合職(製造技術・研究・営業・経理等)", "技術系・事務系総合職/エリア総合職(研究・生産・営業等)", "技術系(開発・設計・生産技術)・営業・管理系"]),
    ("", None),
    ("【OpenWork・働き方】", None),
    ("OpenWork総合評価(5点)", [
        "3.38(109人)", "2.70(18人)", "2.51", "2.90(39人)",
        "3.05(9人)", "2.92(13人)", "2.89(86人)", "2.51(38人)"]),
    ("待遇面の満足度", [
        "3.4", "2.6", "2.75", "2.9",
        "3.0", "3.1", "2.9", "2.3"]),
    ("社員の士気", [
        "2.9", "2.6", "2.7", "—",
        "2.8", "2.7", "2.5", "2.4"]),
    ("20代成長環境", [
        "2.9", "2.2", "—", "—",
        "2.8", "2.5", "2.7", "2.8"]),
    ("平均年収(会社全体)", [
        "日経945万/OW626万(43人)", "日経589万/OW466万", "非公表(30代参考454万)", "OW非公表",
        "日経589万/OW非公表(9人)", "OW506万(10人)/有価証券報告書650万", "日経729万/OW530万(36人)", "日経680万/OW541万(22人)"]),
    ("30歳前後の年収目安", [
        "非公表(商社系・平均年収水準は高め)", "450〜500万円台", "参考30代454万円", "非公表(基本給283,000円+賞与)",
        "非公表(勤続長・年功序列)", "非公表(25-29歳参考値要確認)", "25-29歳平均約433万円(OW参考)", "非公表"]),
    ("有給休暇消化率", [
        "69.0%(OW)/実績14.8日", "58.8%", "低い評価あり/実績8.4日", "56.0%(OW)/実績14.7日",
        "77.2%(OW)/実績14.3日(2025年度)", "69.5%(OW)/実績14.3日(2025年度)", "57.5%(OW)/実績15.5日(2023年度)", "66.4%(OW)/実績13.5日(2024年度)"]),
    ("残業時間(月・マイナビ)", [
        "7.3時間", "9.3時間", "18.7時間", "13.7時間",
        "22.2時間(2025年度)", "7.6時間(2025年度)", "20時間以内", "26.6時間(2024年度)"]),
    ("残業時間(月・OpenWork)", [
        "23.2時間", "12.2時間", "17.7時間", "—",
        "8.2時間", "7.2時間", "21.9時間", "26.5時間"]),
    ("平均勤続年数", [
        "13年", "17年", "11.6年", "16.1年",
        "19.6年(男19.1女21.4)", "19.0年(2025/3)", "16.8年", "16.4年(2024年度)"]),
    ("", None),
    ("【給与・福利厚生】", None),
    ("初任給(大卒・月)", [
        "293,000〜303,000円(理系・首都圏・手当含む)", "237,000円", "251,000円(2025/11引上げ後)", "283,000円(大卒基本給)",
        "234,000円(大卒・2025/4)※院了252,000円", "266,500円(大卒・手当込)※院了287,100円", "287,000円(大卒)※院了309,000円", "273,000円(大卒)※院了288,000円"]),
    ("賞与実績(昨年・年間)", [
        "年2回・口コミ参考約5ヶ月(要確認)", "年2回・昨年実績約4ヶ月分", "年2回・7.2ヶ月分", "年2回・約5.5ヶ月分",
        "年2回", "年2回・5.81ヶ月(2026年予定)", "年2回(6・12月)", "年2回(6・12月)"]),
    ("年間休日", [
        "125日前後(採用サイトFAQ)", "120日前後(本社)", "120日", "120日",
        "125日", "123日", "121日以上(2026年度実績)", "123日"]),
    ("住宅手当・社宅", [
        "社宅・借上社宅(勤務地により)", "独身寮(月1万程度)", "借上社宅", "寮・社宅(独身寮ドリーム広畑)",
        "独身寮・住宅資金貸付制度", "借上げ社宅(東京基準：月6万円補助)", "社宅(三重・愛知工場)・住宅手当", "社宅(転居伴う配属時)・帰省旅費手当"]),
    ("食事補助", [
        "社員食堂(工場)", "食事手当", "採用サイトで言及", "食事補助9,000円/月",
        "要確認", "要確認", "要確認", "食堂(各工場)・食事手当600円/日(工場外)"]),
    ("資格取得支援", [
        "TOEIC支援・英語研修・キャリアアップ支援", "資格・通信教育補助", "消防設備士受験料全額+奨励金", "資格取得報奨金(語学・設備関連)",
        "英会話研修補助・各種階級別研修", "資格奨励金・通信教育補助(70%負担)", "各種研修・eラーニング・ジョブチャレンジ", "資格取得奨励金・通信教育・eラーニング"]),
    ("その他特色制度", [
        "メンター・1on1・健康経営優良法人2026・2030ビジョン", "メンター・社内検定", "水・木ノー残業・19時PC停止", "ノー残業デー・ベースアップ・健康経営優良法人2026",
        "フレックス・時間単位年休・テレワーク可", "フレックス・社員持株会・退職年金", "フレックス・テレワーク・ジョブチャレンジ制度", "リフレッシュ休暇(10年ごと5日)・不妊治療休暇"]),
    ("育児休業取得率(男性)", [
        "64.7%(2024年度)", "55.6%", "100%", "87.5%(2023年度)",
        "62.5%(2025年度)", "75.0%(2025年度)", "28.6%(2023年度)", "35.0%"]),
    ("", None),
    ("【27卒・就活メモ】", None),
    ("事業の軸", [
        "エレクトロニクス商社×ケミカルメーカー(半導体・工業薬品)", "化学コーティング・フィルム・クロス", "防災・消防設備", "有機顔料・色材(液晶・インキ・塗料・トナー)",
        "舶用ディーゼルエンジン・推進装置(内航船55%シェア)", "化成品(高純度リン酸/半導体向け)・機械・電子材料", "電線・ケーブル・光ファイバ・超電導・電力インフラ", "特装車(ゴミ収集車・ダンプ等)・環境・パーキング"]),
    ("マイナビ2027", [
        "掲載あり・オンライン説明会・エントリー受付中", "掲載あり", "掲載あり・説明会で面接確約", "掲載あり・WEB説明会受付中",
        "掲載あり(corp55663)・募集6〜10名", "掲載あり(corp57519)・募集11〜15名", "掲載あり(corp85245)", "掲載あり(corp2639)・募集25〜30名"]),
    ("採用実績に大阪電気通信大学(マイナビ2027)", [
        "なし(電気通信大学はあり)", "あり(大学・採用実績学校)", "あり(大学)", "なし",
        "あり(大学・採用実績校に明記)", "なし(掲載なし)", "あり(2025年度採用実績に記載)", "あり(大阪電気通信大学)"]),
    ("データ出典・更新", [
        "マイナビ2027/OpenWork/日経(2026年5月)", "同上", "同上", "同上",
        "マイナビ2027/OpenWork/公式採用サイト(2026年6月)", "同上", "同上", "同上"]),
]

# ざっくり比較: 値は文字列1つ（全カラム結合表示）
AXIS_ROWS = [
    ("【ざっくり比較（就活の軸）】", None),
    ("初任給（大卒・月給）",
     "伯東(29.3〜30.3万) ＞ SWCC(28.7万) ＞ 山陽色素(28.3万) ＞ 極東開発(27.3万) ＞ ラサ工業(26.65万) ＞ ヤマトプロテック(25.1万) ＞ ダイニック(23.7万) ＞ 阪神内燃(23.4万)"),
    ("資本金",
     "SWCC(242億) ＞ 極東開発(119億) ＞ ラサ工業(84億) ＞ 伯東(81億) ＞ ダイニック(57億) ＞ 山陽色素・ヤマトプロテック・阪神内燃(〜9億)"),
    ("会社規模・売上",
     "SWCC(2,378億) ＞ 伯東(1,831億) ＞ 極東開発(1,404億) ＞ ラサ工業(477億) ＞ ダイニック(440億) ＞ ヤマトプロテック(360億) ＞ 山陽色素(240億) ＞ 阪神内燃(140億)"),
    ("口コミ総合（OpenWork）",
     "伯東(3.38) ＞ 阪神内燃(3.05) ＞ ラサ工業(2.92) ＞ 山陽色素(2.90) ＞ SWCC(2.89) ＞ ダイニック(2.70) ＞ 極東開発=ヤマトプロテック(2.51)"),
    ("残業の少なさ（マイナビ実績）",
     "ラサ工業(7.6h) ＞ 伯東(7.3h) ＞ ダイニック(9.3h) ＞ 山陽色素(13.7h) ＞ ヤマトプロテック(18.7h) ＞ SWCC(20h以内) ＞ 阪神内燃(22.2h) ＞ 極東開発(26.6h)"),
    ("賞与（昨年実績・年間）",
     "ヤマトプロテック(7.2ヶ月) ＞ ラサ工業(5.81ヶ月予定) ＞ 山陽色素(5.5ヶ月) ＞ 伯東・ダイニック(4〜5ヶ月前後) ／ SWCC・極東開発・阪神内燃は年2回(実績月数要確認)"),
    ("理系・化学系の専門性",
     "ラサ工業(半導体向けリン酸・化学)／伯東(商社×化学)／山陽色素(有機顔料)／ダイニック(コーティング)／SWCC(電線・超電導)／極東開発(特装車・機械)／ヤマト(防災)／阪神内燃(舶用エンジン)"),
    ("社会貢献・インフラ感",
     "SWCC(電力インフラ)・極東開発(廃棄物処理・特装)・ヤマトプロテック(防災)・阪神内燃(海運インフラ) ＞ 伯東・ラサ工業(半導体) ＞ 山陽色素・ダイニック"),
    ("上場・将来性（株価）",
     "プライム: 伯東・ラサ工業(半導体需要◎)・SWCC・極東開発 ／ スタンダード: ダイニック・阪神内燃機工業 ／ 非上場: ヤマトプロテック・山陽色素"),
    ("", None),
    ("出典: マイナビ2027 / OpenWork / 公式採用サイト / 日経会社情報・決算公告等。2026年6月時点。",
     "賞与・初任給の一部はユーザー確認値を反映。"),
]

URLS = [
    ("伯東",           "https://job.mynavi.jp/27/pc/search/corp290/outline.html"),
    ("ダイニック",     "https://job.mynavi.jp/27/pc/search/corp1234/outline.html"),
    ("ヤマトプロテック","https://job.mynavi.jp/27/"),
    ("山陽色素",       "https://job.mynavi.jp/27/"),
    ("阪神内燃機工業", "https://job.mynavi.jp/27/pc/search/corp55663/outline.html"),
    ("ラサ工業",       "https://job.mynavi.jp/27/pc/search/corp57519/outline.html"),
    ("SWCC",           "https://job.mynavi.jp/27/pc/search/corp85245/outline.html"),
    ("極東開発工業",   "https://job.mynavi.jp/27/pc/search/corp2639/outline.html"),
]

# ============================================================
# Excel 生成
# ============================================================
HC  = "1F4E79"   # ヘッダー濃紺
SC  = "2E75B6"   # セクション中青
ALT = "DEEAF1"   # 交互薄青
W   = "FFFFFF"

def border():
    s = Side(style='thin', color='AAAAAA')
    return Border(left=s, right=s, top=s, bottom=s)

def write_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_MAIN
    ws.column_dimensions['A'].width = 26
    for i in range(2, 2 + N):
        ws.column_dimensions[get_column_letter(i)].width = 22

    # ヘッダー
    ws.row_dimensions[1].height = 26
    for col, val in enumerate(["比較項目"] + CN, 1):
        c = ws.cell(row=1, column=col, value=val)
        c.font = Font(name='Meiryo', bold=True, color=W, size=10)
        c.fill = PatternFill("solid", fgColor=HC)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border()

    all_rows = ROWS + AXIS_ROWS
    ridx = 2
    alt = False
    for label, values in all_rows:
        is_sec  = label.startswith('【')
        is_emp  = label == ''
        is_axis = isinstance(values, str)  # ざっくり比較の値行
        is_foot = is_axis and label.startswith('出典')

        if is_sec:
            bg = SC
        elif is_emp:
            bg = W
        else:
            bg = ALT if alt else W
            alt = not alt

        ws.row_dimensions[ridx].height = 36 if (is_sec or is_axis) else 30

        a = ws.cell(row=ridx, column=1, value=label)
        a.font = Font(name='Meiryo', bold=is_sec or is_foot, color=(W if is_sec else "000000"), size=9)
        a.fill = PatternFill("solid", fgColor=bg)
        a.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        a.border = border()

        if values is None:
            # セクション見出しまたは空行: B〜I も同色で塗る
            for col in range(2, 2 + N):
                c = ws.cell(row=ridx, column=col, value="")
                c.fill = PatternFill("solid", fgColor=bg)
                c.border = border()
        elif is_axis:
            # ざっくり比較: B2〜I2 をマージして値を入れる
            ws.merge_cells(start_row=ridx, start_column=2, end_row=ridx, end_column=1 + N)
            c = ws.cell(row=ridx, column=2, value=values)
            c.font = Font(name='Meiryo', size=8.5)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            c.border = border()
        else:
            # 通常データ行
            for i, val in enumerate(values):
                c = ws.cell(row=ridx, column=i + 2, value=val)
                c.font = Font(name='Meiryo', size=8.5)
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = border()
        ridx += 1

    # 参照URLシート
    ws2 = wb.create_sheet(SHEET_URLS)
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 80
    ws2.append(["会社名", "マイナビ2027 URL"])
    for c in ws2[1]:
        c.font = Font(bold=True, color=W)
        c.fill = PatternFill("solid", fgColor=HC)
        c.alignment = Alignment(horizontal='center')
    for row in URLS:
        ws2.append(list(row))

    wb.save(OUT_XLSX)
    print(f"Excel saved: {OUT_XLSX}")

# ============================================================
# PDF 生成 (A4横1枚・行高さ自動 + 自動スケーリング)
# ============================================================
def jp(text, fs=4.8, col=colors.black, align='CENTER'):
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    return Paragraph(str(text) if text else '', ParagraphStyle(
        'j', fontName=JP, fontSize=fs, leading=fs * 1.35,
        textColor=col,
        alignment=TA_LEFT if align == 'LEFT' else TA_CENTER,
        wordWrap='CJK',
        spaceAfter=0, spaceBefore=0))

def build_table(fs_hdr=5.5, fs_sec=5.0, fs_lbl=4.6, fs_val=4.3, fs_ax=4.3):
    """テーブルデータと行高さ(None=自動)を返す"""
    PW, PH = landscape(A4)
    ML = MR = 8
    cw = PW - ML - MR
    LW = 82
    DW = (cw - LW) / N

    HC_c  = colors.HexColor('#1F4E79')
    SC_c  = colors.HexColor('#2E75B6')
    ALT_c = colors.HexColor('#DEEAF1')
    W_c   = colors.white
    BK    = colors.black

    table_data  = []
    row_styles  = []
    span_cmds   = []

    # ヘッダー行
    table_data.append(
        [jp("比較項目", fs=fs_hdr, col=W_c, align='LEFT')] +
        [jp(c, fs=fs_hdr, col=W_c) for c in CN]
    )
    row_styles.append(('BACKGROUND', (0,0), (-1,0), HC_c))

    ridx = 1
    alt  = False
    for label, values in ROWS + AXIS_ROWS:
        is_sec  = label.startswith('【')
        is_emp  = label == ''
        is_axis = isinstance(values, str)

        if is_sec:
            bg = SC_c
        elif is_emp:
            bg = W_c
        else:
            bg = ALT_c if alt else W_c
            alt = not alt

        lc = W_c if is_sec else BK

        if values is None:
            row = [jp(label, fs=fs_sec if is_sec else 3.5, col=lc, align='LEFT')] + \
                  [jp('') for _ in CN]
            if not is_emp:
                span_cmds.append(('SPAN', (0, ridx), (N, ridx)))
        elif is_axis:
            row = [jp(label, fs=fs_lbl, col=BK, align='LEFT'),
                   jp(values, fs=fs_ax, col=BK, align='LEFT')] + \
                  [jp('') for _ in range(N-1)]
            span_cmds.append(('SPAN', (1, ridx), (N, ridx)))
        else:
            row = [jp(label, fs=fs_sec if is_sec else fs_lbl, col=lc, align='LEFT')] + \
                  [jp(v, fs=fs_val, col=(W_c if is_sec else BK)) for v in values]

        table_data.append(row)
        row_styles.append(('BACKGROUND', (0, ridx), (-1, ridx), bg))
        if is_sec:
            row_styles.append(('TEXTCOLOR', (0, ridx), (-1, ridx), W_c))
        ridx += 1

    col_widths = [LW] + [DW] * N
    base = [
        ('FONTNAME',      (0,0),(-1,-1), JP),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
        ('ALIGN',         (0,0),(-1,-1), 'CENTER'),
        ('ALIGN',         (0,0),(0,-1),  'LEFT'),
        ('GRID',          (0,0),(-1,-1), 0.25, colors.HexColor('#AAAAAA')),
        ('LEFTPADDING',   (0,0),(-1,-1), 2),
        ('RIGHTPADDING',  (0,0),(-1,-1), 2),
        ('TOPPADDING',    (0,0),(-1,-1), 2),
        ('BOTTOMPADDING', (0,0),(-1,-1), 2),
    ]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle(base + row_styles + span_cmds))
    return t, col_widths

def write_pdf():
    from reportlab.pdfgen import canvas as cv_mod

    PW, PH = landscape(A4)
    ML = MR = 8; MT = MB = 10
    avail_w = PW - ML - MR
    avail_h = PH - MT - MB

    # ① 自動行高さでテーブルを作成
    t, col_widths = build_table()

    # ② 実際の描画サイズを測定
    tmp_cv = cv_mod.Canvas(TMP_PDF, pagesize=landscape(A4))
    nat_w, nat_h = t.wrapOn(tmp_cv, avail_w, 99999)
    print(f"  Table natural size: {nat_w:.1f} x {nat_h:.1f}  avail: {avail_w:.1f} x {avail_h:.1f}")

    # ③ スケール計算（収まる場合は 1.0）
    scale = min(1.0, avail_h / nat_h, avail_w / nat_w)
    print(f"  Scale: {scale:.4f}")

    # ④ キャンバスに直接描画
    c = cv_mod.Canvas(TMP_PDF, pagesize=landscape(A4))
    c.saveState()
    # 左下マージンを起点にスケーリング後配置
    c.translate(ML, MB + avail_h - nat_h * scale)
    c.scale(scale, scale)
    t.drawOn(c, 0, 0)
    c.restoreState()
    c.save()

    shutil.copy2(TMP_PDF, OUT_PDF)
    print(f"PDF saved: {OUT_PDF}")

if __name__ == '__main__':
    write_xlsx()
    write_pdf()
    print("Done.")
